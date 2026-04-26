"""Export BOQ revisi ke Excel dengan formula dinamis."""
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import BOQItem, BOQRevision

THIN = Side(style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _money_format() -> str:
    # 1.234.567,89 (pemisah ribuan titik, desimal koma — Indonesia)
    return '_-"Rp"\\ #.##0,00_-;[Red]_-"Rp"\\ #.##0,00\\-_-;_-"Rp"\\ "-"??_-'


def _num_format() -> str:
    return "#,##0.0000"


def export_revision_xlsx(revision: BOQRevision) -> bytes:
    contract = revision.contract
    wb = Workbook()
    ws = wb.active
    ws.title = f"BOQ V{revision.version}"

    # Header info kontrak (Bagian 14.2)
    ws["A1"] = "BILL OF QUANTITY (BOQ)"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:I1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Nomor Kontrak"
    ws["B3"] = contract.number
    ws["A4"] = "Nama Kontrak"
    ws["B4"] = contract.name
    ws["A5"] = "PPK"
    ws["B5"] = contract.ppk.full_name if contract.ppk_id else "-"
    ws["A6"] = "Kontraktor"
    ws["B6"] = contract.contractor.name if contract.contractor_id else "-"
    ws["A7"] = "Tahun Anggaran"
    ws["B7"] = contract.fiscal_year
    ws["A8"] = "Revisi"
    ws["B8"] = f"V{revision.version} - {revision.get_status_display()}"
    ws["A9"] = "PPN"
    ws["B9"] = f"{contract.ppn_pct}%"
    ws["A10"] = "Periode"
    ws["B10"] = f"{contract.start_date} s.d. {contract.end_date}"

    for r in range(3, 11):
        ws[f"A{r}"].font = Font(bold=True)

    header_row = 12
    headers = [
        "Kode", "Uraian Pekerjaan", "Fasilitas", "Sat.",
        "Volume", "Harga Satuan (PRE-PPN)", "Jumlah", "Bobot %",
        "Lvl",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    # Order items: by facility, lalu display_order, lalu full_code
    items = list(
        BOQItem.objects.filter(boq_revision=revision)
                          .select_related("facility")
                          .order_by("facility__display_order", "facility__code",
                                      "display_order", "full_code")
    )

    cur_row = header_row + 1
    for it in items:
        # Indent uraian by level
        indent = "  " * it.level
        ws.cell(row=cur_row, column=1, value=it.full_code or it.code)
        ws.cell(row=cur_row, column=2, value=indent + it.description)
        ws.cell(row=cur_row, column=3, value=f"{it.facility.code} - {it.facility.name}")
        ws.cell(row=cur_row, column=4, value=it.unit)
        if it.is_leaf:
            ws.cell(row=cur_row, column=5, value=float(it.volume))
            ws.cell(row=cur_row, column=6, value=float(it.unit_price))
            # Formula: jumlah = volume × harga_satuan
            ws.cell(row=cur_row, column=7, value=f"=E{cur_row}*F{cur_row}")
        else:
            # Parent: total = SUMnya (will be set after we know which child rows)
            # Simpan placeholder; parent total dihitung server-side akan diatur lewat formula sederhana.
            # Lebih praktis: tulis nilai langsung (server-side recompute) supaya hasil deterministik.
            ws.cell(row=cur_row, column=7, value=float(it.total_price))
            for col in (5, 6):
                ws.cell(row=cur_row, column=col, value="")

        if parse_pct := float(it.weight_pct):
            ws.cell(row=cur_row, column=8, value=parse_pct / 100.0)
        else:
            ws.cell(row=cur_row, column=8, value="")

        ws.cell(row=cur_row, column=9, value=it.level)

        # Style berdasarkan leaf
        for col in range(1, 10):
            cell = ws.cell(row=cur_row, column=col)
            cell.border = BORDER
            if not it.is_leaf:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="EFEFEF")

        # Format kolom
        ws.cell(row=cur_row, column=5).number_format = _num_format()
        ws.cell(row=cur_row, column=6).number_format = _money_format()
        ws.cell(row=cur_row, column=7).number_format = _money_format()
        ws.cell(row=cur_row, column=8).number_format = "0.00%"

        cur_row += 1

    # Footer total
    total_leaf = (
        BOQItem.objects.filter(boq_revision=revision, is_leaf=True)
                          .aggregate(t=Sum("total_price"))["t"] or Decimal("0")
    )
    ws.cell(row=cur_row, column=1, value="TOTAL (PRE-PPN)").font = Font(bold=True)
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
    cell = ws.cell(row=cur_row, column=7, value=float(total_leaf))
    cell.font = Font(bold=True)
    cell.number_format = _money_format()
    for col in range(1, 10):
        ws.cell(row=cur_row, column=col).border = BORDER
    cur_row += 1

    # PPN
    ppn_amount = (total_leaf * contract.ppn_pct / Decimal("100")).quantize(Decimal("0.01"))
    ws.cell(row=cur_row, column=1, value=f"PPN ({contract.ppn_pct}%)")
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
    ws.cell(row=cur_row, column=7, value=float(ppn_amount)).number_format = _money_format()
    cur_row += 1

    # Total post-PPN
    total_post = total_leaf + ppn_amount
    ws.cell(row=cur_row, column=1, value="TOTAL (POST-PPN)").font = Font(bold=True)
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
    cell = ws.cell(row=cur_row, column=7, value=float(total_post))
    cell.font = Font(bold=True, color="1F4E78", size=12)
    cell.number_format = _money_format()
    for col in range(1, 10):
        ws.cell(row=cur_row, column=col).border = BORDER

    # Auto width
    widths = [14, 60, 28, 8, 14, 22, 22, 10, 6]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
