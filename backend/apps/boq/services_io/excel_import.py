"""Import BOQ dari Excel — Format A (simple template).

Format A (CLAUDE.md 11.1):
    facility_code, facility_name, code, parent_code, description, unit,
    volume, unit_price, total_price, planned_start_week, planned_duration_weeks

Aturan:
- facility_code, description, volume wajib (volume boleh 0 utk lumpsum).
- parent_code dipakai utk hirarki — chain kode parent.
- Level di-derive otomatis (max 0..3).
- Code unik per fasilitas dalam revisi.
- Import tidak boleh overwrite Nilai Kontrak.

Hasil:
- preview_only=True: return rencana (counts, errors, mappings) tanpa commit.
- preview_only=False: lakukan replace seluruh items revisi (atomic).
"""
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import IO

from django.db import transaction
from openpyxl import load_workbook

from apps.contract.models import Facility
from common.exceptions import DomainError

from ..models import BOQItem, BOQRevision
from ..services import assert_revision_writable, recompute_all

REQUIRED_HEADERS = [
    "facility_code", "code", "description",
]
OPTIONAL_HEADERS = [
    "facility_name", "parent_code", "unit",
    "volume", "unit_price", "total_price",
    "planned_start_week", "planned_duration_weeks",
]
ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS


@dataclass
class ImportRow:
    row_number: int
    facility_code: str
    facility_name: str = ""
    code: str = ""
    parent_code: str = ""
    description: str = ""
    unit: str = ""
    volume: Decimal = Decimal("0")
    unit_price: Decimal = Decimal("0")
    planned_start_week: int | None = None
    planned_duration_weeks: int | None = None
    errors: list[str] = field(default_factory=list)


@dataclass
class ImportPreview:
    detected_format: str = "A"
    sheet_used: str = ""
    rows_total: int = 0
    rows_valid: int = 0
    rows_invalid: int = 0
    facility_summary: list[dict] = field(default_factory=list)
    unmatched_facility_codes: list[str] = field(default_factory=list)
    sample_errors: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _to_decimal(val, default=Decimal("0")) -> Decimal:
    if val is None or val == "":
        return default
    try:
        return Decimal(str(val).replace(",", ".").strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"Bukan angka valid: {val!r}")


def _to_int(val) -> int | None:
    if val is None or val == "":
        return None
    try:
        return int(val)
    except (TypeError, ValueError):
        raise ValueError(f"Bukan integer valid: {val!r}")


def _normalize_header(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "_")


def _detect_format_and_sheet(wb) -> tuple[str, str]:
    """Cari sheet pertama yang punya header Format A. Return (format, sheet_name)."""
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            continue
        headers = [_normalize_header(c or "") for c in first_row]
        # Cek minimal punya facility_code + code + description
        if all(h in headers for h in REQUIRED_HEADERS):
            return "A", ws.title
    raise DomainError(
        "Format Excel tidak dikenali. Minimal sheet pertama harus punya kolom: "
        + ", ".join(REQUIRED_HEADERS) + ".",
        code="UNRECOGNIZED_FORMAT",
    )


def _parse_rows_format_a(ws) -> list[ImportRow]:
    headers = [_normalize_header(c or "") for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(headers)}

    def get(row, key, default=None):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else default

    rows: list[ImportRow] = []
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip baris kosong
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        ir = ImportRow(row_number=n, facility_code=str(get(row, "facility_code") or "").strip())
        ir.facility_name = str(get(row, "facility_name") or "").strip()
        ir.code = str(get(row, "code") or "").strip()
        ir.parent_code = str(get(row, "parent_code") or "").strip()
        ir.description = str(get(row, "description") or "").strip()
        ir.unit = str(get(row, "unit") or "").strip()
        try:
            ir.volume = _to_decimal(get(row, "volume"))
        except ValueError as e:
            ir.errors.append(f"volume: {e}")
        try:
            ir.unit_price = _to_decimal(get(row, "unit_price"))
        except ValueError as e:
            ir.errors.append(f"unit_price: {e}")
        try:
            ir.planned_start_week = _to_int(get(row, "planned_start_week"))
        except ValueError as e:
            ir.errors.append(f"planned_start_week: {e}")
        try:
            ir.planned_duration_weeks = _to_int(get(row, "planned_duration_weeks"))
        except ValueError as e:
            ir.errors.append(f"planned_duration_weeks: {e}")

        # Required field validation
        if not ir.facility_code:
            ir.errors.append("facility_code wajib diisi.")
        if not ir.code:
            ir.errors.append("code wajib diisi.")
        if not ir.description:
            ir.errors.append("description wajib diisi.")
        if ir.volume < 0:
            ir.errors.append("volume tidak boleh negatif.")
        if ir.unit_price < 0:
            ir.errors.append("unit_price tidak boleh negatif.")

        rows.append(ir)
    return rows


def parse_excel(file: IO) -> tuple[list[ImportRow], ImportPreview]:
    wb = load_workbook(filename=file, read_only=True, data_only=True)
    fmt, sheet_name = _detect_format_and_sheet(wb)
    ws = wb[sheet_name]
    rows = _parse_rows_format_a(ws)

    preview = ImportPreview(
        detected_format=fmt, sheet_used=sheet_name,
        rows_total=len(rows),
        rows_valid=sum(1 for r in rows if not r.errors),
        rows_invalid=sum(1 for r in rows if r.errors),
    )
    # Sample errors (max 20)
    for r in rows:
        if r.errors:
            preview.sample_errors.append({
                "row": r.row_number, "code": r.code, "errors": r.errors,
            })
            if len(preview.sample_errors) >= 20:
                break

    # Facility summary
    by_fac: dict[str, dict] = {}
    for r in rows:
        if not r.facility_code:
            continue
        d = by_fac.setdefault(r.facility_code, {
            "facility_code": r.facility_code,
            "facility_name": r.facility_name,
            "row_count": 0,
            "valid_count": 0,
        })
        d["row_count"] += 1
        if not r.errors:
            d["valid_count"] += 1
        if r.facility_name and not d["facility_name"]:
            d["facility_name"] = r.facility_name
    preview.facility_summary = list(by_fac.values())
    return rows, preview


def check_facility_mapping(rows: list[ImportRow], revision: BOQRevision) -> list[str]:
    """Cari facility_code yang tidak match Facility existing di kontrak."""
    fac_codes = {r.facility_code for r in rows if r.facility_code}
    existing = set(
        Facility.objects.filter(
            location__contract=revision.contract,
            deleted_at__isnull=True,
        ).values_list("code", flat=True)
    )
    return sorted(fac_codes - existing)


@transaction.atomic
def commit_import(
    rows: list[ImportRow], revision: BOQRevision, *, replace_existing: bool = True,
) -> dict:
    """Commit hasil parse ke DB. Atomic.

    Strategi:
    - Resolve facility_code -> Facility.id (via contract scope)
    - Build map kode -> tentative item
    - 2-pass insert: root dulu, lalu child by parent_code
    """
    assert_revision_writable(revision)

    # Filter rows yang valid
    valid_rows = [r for r in rows if not r.errors]
    if not valid_rows:
        raise DomainError("Tidak ada baris valid untuk di-import.", code="NO_VALID_ROWS")

    # Resolve facility map
    fac_codes = {r.facility_code for r in valid_rows}
    fac_map: dict[str, Facility] = {
        f.code: f for f in Facility.objects.filter(
            location__contract=revision.contract,
            code__in=fac_codes, deleted_at__isnull=True,
        ).select_related("location")
    }
    missing = fac_codes - set(fac_map.keys())
    if missing:
        raise DomainError(
            f"facility_code tidak ditemukan di kontrak: {', '.join(sorted(missing))}. "
            "Buat fasilitas terlebih dahulu di tab Lokasi.",
            code="UNMATCHED_FACILITY",
        )

    if replace_existing:
        BOQItem.objects.filter(boq_revision=revision).delete()

    # Pass 1: insert root (parent_code kosong) per fasilitas
    by_fac_code_map: dict[tuple[str, str], BOQItem] = {}  # (facility_code, code) -> item
    pending: list[ImportRow] = []
    for r in valid_rows:
        if r.parent_code:
            pending.append(r)
            continue
        fac = fac_map[r.facility_code]
        item = BOQItem.objects.create(
            boq_revision=revision,
            facility=fac,
            code=r.code,
            description=r.description,
            unit=r.unit,
            volume=r.volume,
            unit_price=r.unit_price,
            parent=None,
            display_order=r.row_number,
            planned_start_week=r.planned_start_week,
            planned_duration_weeks=r.planned_duration_weeks,
        )
        by_fac_code_map[(r.facility_code, r.code)] = item

    # Pass 2-N: insert berulang sampai semua pending masuk (max iter = depth)
    last_pending = -1
    iter_n = 0
    while pending and len(pending) != last_pending and iter_n < 8:
        last_pending = len(pending)
        iter_n += 1
        new_pending: list[ImportRow] = []
        for r in pending:
            parent = by_fac_code_map.get((r.facility_code, r.parent_code))
            if not parent:
                new_pending.append(r)
                continue
            fac = fac_map[r.facility_code]
            item = BOQItem.objects.create(
                boq_revision=revision,
                facility=fac,
                code=r.code,
                description=r.description,
                unit=r.unit,
                volume=r.volume,
                unit_price=r.unit_price,
                parent=parent,
                display_order=r.row_number,
                planned_start_week=r.planned_start_week,
                planned_duration_weeks=r.planned_duration_weeks,
            )
            by_fac_code_map[(r.facility_code, r.code)] = item
        pending = new_pending

    if pending:
        codes = [(r.facility_code, r.code, r.parent_code) for r in pending[:5]]
        raise DomainError(
            f"Parent tidak ditemukan utk {len(pending)} baris. Contoh: {codes}",
            code="ORPHAN_ITEMS",
        )

    # Recompute level/full_code/leaf-flag/total/weight
    recompute_all(revision)

    return {
        "rows_imported": len(valid_rows),
        "rows_skipped": sum(1 for r in rows if r.errors),
        "facilities_used": len(fac_map),
    }


def build_template_workbook():
    """Generate workbook template kosong dengan header + 1 baris contoh."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    headers = [
        "facility_code", "facility_name", "code", "parent_code",
        "description", "unit", "volume", "unit_price", "total_price",
        "planned_start_week", "planned_duration_weeks",
    ]
    ws.append(headers)
    # Header style
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample rows
    samples = [
        ["F1", "Gudang Beku 1", "1", "", "Pekerjaan Persiapan", "ls", 1, 5_000_000, "", 1, 2],
        ["F1", "Gudang Beku 1", "1.1", "1", "Mobilisasi", "ls", 1, 3_000_000, "", 1, 1],
        ["F1", "Gudang Beku 1", "2", "", "Pekerjaan Struktur", "", 0, 0, "", 3, 8],
        ["F1", "Gudang Beku 1", "2.1", "2", "Beton K-300", "m3", 50, 1_500_000, "", 3, 4],
    ]
    for s in samples:
        ws.append(s)

    # Auto width sederhana
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Sheet README
    ws2 = wb.create_sheet("README")
    ws2["A1"] = "Template Import BOQ - Format A"
    ws2["A1"].font = Font(bold=True, size=14)
    notes = [
        "",
        "Cara pakai:",
        "1. Isi sheet 'BOQ' dengan baris-baris item BOQ Anda. Hapus baris contoh.",
        "2. facility_code WAJIB diisi dan harus sama dengan kode Fasilitas yang sudah",
        "   dibuat di tab Lokasi & Fasilitas pada kontrak.",
        "3. parent_code = kode item parent (jika item ini sub-pekerjaan). Kosongkan",
        "   untuk root.",
        "4. volume boleh 0 untuk item lumpsum. unit_price PRE-PPN (tanpa PPN).",
        "5. total_price akan dihitung otomatis sistem (volume × unit_price).",
        "6. Setelah upload, sistem akan PREVIEW data terlebih dahulu — Anda bisa",
        "   batal sebelum benar-benar di-commit ke database.",
    ]
    for i, n in enumerate(notes, start=2):
        ws2[f"A{i}"] = n
    ws2.column_dimensions["A"].width = 80

    return wb
