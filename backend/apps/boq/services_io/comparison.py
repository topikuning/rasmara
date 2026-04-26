"""Komparasi antar revisi BOQ (Bagian 11.2 + 14.4)."""
from dataclasses import dataclass, field
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import BOQItem, BOQRevision


@dataclass
class CompareLine:
    full_code: str = ""
    description: str = ""
    facility_code: str = ""
    unit: str = ""
    unit_price_a: Decimal = Decimal("0")
    unit_price_b: Decimal = Decimal("0")
    volume_a: Decimal = Decimal("0")
    volume_b: Decimal = Decimal("0")
    total_a: Decimal = Decimal("0")
    total_b: Decimal = Decimal("0")
    diff_volume: Decimal = Decimal("0")     # B - A
    diff_total: Decimal = Decimal("0")       # B - A
    note: str = ""


@dataclass
class CompareResult:
    revision_a: dict = field(default_factory=dict)
    revision_b: dict = field(default_factory=dict)
    lines: list[CompareLine] = field(default_factory=list)
    total_a: Decimal = Decimal("0")
    total_b: Decimal = Decimal("0")
    total_tambah: Decimal = Decimal("0")
    total_kurang: Decimal = Decimal("0")


def _key(item: BOQItem) -> str:
    """Identitas item untuk matching antar revisi.

    Pakai (facility_code, full_code) — stabil walau parent berubah.
    """
    fac_code = item.facility.code if item.facility_id else ""
    return f"{fac_code}::{item.full_code or item.code}"


def compare_revisions(rev_a: BOQRevision, rev_b: BOQRevision) -> CompareResult:
    """Compare A vs B (B = revisi target/baru, A = revisi referensi/lama)."""
    if rev_a.contract_id != rev_b.contract_id:
        from common.exceptions import DomainError
        raise DomainError("Revisi A dan B harus dari kontrak yang sama.",
                          code="CROSS_CONTRACT_COMPARE")

    items_a = {
        _key(it): it
        for it in BOQItem.objects.filter(boq_revision=rev_a, is_leaf=True)
                                    .select_related("facility")
    }
    items_b = {
        _key(it): it
        for it in BOQItem.objects.filter(boq_revision=rev_b, is_leaf=True)
                                    .select_related("facility")
    }

    keys = sorted(set(items_a.keys()) | set(items_b.keys()))
    lines: list[CompareLine] = []

    for k in keys:
        a = items_a.get(k)
        b = items_b.get(k)
        ref = b or a
        line = CompareLine(
            full_code=(ref.full_code or ref.code) if ref else "",
            description=ref.description if ref else "",
            facility_code=ref.facility.code if ref else "",
            unit=ref.unit if ref else "",
        )
        if a:
            line.volume_a = a.volume
            line.unit_price_a = a.unit_price
            line.total_a = a.total_price
        if b:
            line.volume_b = b.volume
            line.unit_price_b = b.unit_price
            line.total_b = b.total_price

        line.diff_volume = line.volume_b - line.volume_a
        line.diff_total = line.total_b - line.total_a

        if a is None and b is not None:
            line.note = "BARU"
        elif a is not None and b is None:
            line.note = "DIHAPUS"
        elif line.diff_total > 0:
            line.note = "BERTAMBAH"
        elif line.diff_total < 0:
            line.note = "BERKURANG"
        else:
            line.note = "SAMA"
        lines.append(line)

    total_a = (BOQItem.objects.filter(boq_revision=rev_a, is_leaf=True)
                                  .aggregate(t=Sum("total_price"))["t"] or Decimal("0"))
    total_b = (BOQItem.objects.filter(boq_revision=rev_b, is_leaf=True)
                                  .aggregate(t=Sum("total_price"))["t"] or Decimal("0"))
    total_tambah = sum((line.diff_total for line in lines if line.diff_total > 0), Decimal("0"))
    total_kurang = sum((-line.diff_total for line in lines if line.diff_total < 0), Decimal("0"))

    return CompareResult(
        revision_a={"id": str(rev_a.id), "version": rev_a.version,
                     "status": rev_a.status},
        revision_b={"id": str(rev_b.id), "version": rev_b.version,
                     "status": rev_b.status},
        lines=lines, total_a=total_a, total_b=total_b,
        total_tambah=total_tambah, total_kurang=total_kurang,
    )


def export_compare_xlsx(result: CompareResult, contract) -> bytes:
    """Excel komparasi dengan kolom: Jenis Pekerjaan | Harga Satuan |
    Pekerjaan A (Vol & Jumlah) | Pekerjaan B (Vol & Jumlah) |
    Tambah (Vol & Jumlah) | Kurang (Vol & Jumlah) | Ket."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Komparasi BOQ"

    # Header info
    ws["A1"] = f"KOMPARASI BOQ — V{result.revision_a['version']} vs V{result.revision_b['version']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:K1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Kontrak"
    ws["B3"] = f"{contract.number} - {contract.name}"

    header_row = 5
    headers = [
        "Kode", "Jenis Pekerjaan", "Sat.", "Harga Satuan",
        "Vol. A (V" + str(result.revision_a['version']) + ")",
        "Jumlah A",
        "Vol. B (V" + str(result.revision_b['version']) + ")",
        "Jumlah B",
        "Tambah (Vol)", "Tambah (Jumlah)",
        "Ket.",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    money_fmt = '_-"Rp"\\ #.##0,00_-'
    num_fmt = "#,##0.0000"

    cur = header_row + 1
    for line in result.lines:
        ws.cell(row=cur, column=1, value=line.full_code)
        ws.cell(row=cur, column=2, value=line.description)
        ws.cell(row=cur, column=3, value=line.unit)
        ws.cell(row=cur, column=4, value=float(line.unit_price_b or line.unit_price_a))
        ws.cell(row=cur, column=4).number_format = money_fmt

        ws.cell(row=cur, column=5, value=float(line.volume_a))
        ws.cell(row=cur, column=5).number_format = num_fmt
        # Formula jumlah A = vol A * harga
        ws.cell(row=cur, column=6, value=f"=E{cur}*D{cur}")
        ws.cell(row=cur, column=6).number_format = money_fmt

        ws.cell(row=cur, column=7, value=float(line.volume_b))
        ws.cell(row=cur, column=7).number_format = num_fmt
        ws.cell(row=cur, column=8, value=f"=G{cur}*D{cur}")
        ws.cell(row=cur, column=8).number_format = money_fmt

        # Tambah/Kurang
        ws.cell(row=cur, column=9, value=f"=G{cur}-E{cur}")
        ws.cell(row=cur, column=9).number_format = num_fmt
        ws.cell(row=cur, column=10, value=f"=H{cur}-F{cur}")
        ws.cell(row=cur, column=10).number_format = money_fmt

        ws.cell(row=cur, column=11, value=line.note)
        cur += 1

    # Totals
    cur += 1
    ws.cell(row=cur, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=cur, column=6, value=f"=SUM(F{header_row + 1}:F{cur - 2})").number_format = money_fmt
    ws.cell(row=cur, column=8, value=f"=SUM(H{header_row + 1}:H{cur - 2})").number_format = money_fmt
    ws.cell(row=cur, column=10, value=f"=SUM(J{header_row + 1}:J{cur - 2})").number_format = money_fmt
    for col in (1, 6, 8, 10):
        ws.cell(row=cur, column=col).font = Font(bold=True)

    widths = [12, 50, 8, 16, 12, 16, 12, 16, 12, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
